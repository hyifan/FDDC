# -- coding: utf-8 --
import numpy as np
from numpy import *
from openpyxl import load_workbook
from openpyxl import Workbook

def load_training_data():
    # 读取excel表
    wb, sheet_income, sheet_cash, sheet_balance, wb_income, wb_cash, wb_balance, sheet_income_bank, sheet_cash_bank, sheet_balance_bank, sheet_income_insurance, sheet_cash_insurance, sheet_balance_insurance, sheet_income_securities, sheet_cash_securities, sheet_balance_securities = load_xlsx()

    # 读取excel表中数据
    data_income, data_cash, data_balance, data_income_bank, data_cash_bank, data_balance_bank, data_income_insurance, data_cash_insurance, data_balance_insurance, data_income_securities, data_cash_securities, data_balance_securities = load_data(sheet_income, sheet_cash, sheet_balance, sheet_income_bank, sheet_cash_bank, sheet_balance_bank, sheet_income_insurance, sheet_cash_insurance, sheet_balance_insurance, sheet_income_securities, sheet_cash_securities, sheet_balance_securities)

    data_market, data_macro, data_com = load_basic_data(wb)

    # 将数据按各季度拼接
    data_bs, labels_bs, answer_data_bs, data_fa, labels_fa, answer_data_fa, symbol_bs, symbol_fa = load_data_wrapper(data_income, data_cash, data_balance, data_income_bank, data_cash_bank, data_balance_bank, data_income_insurance, data_cash_insurance, data_balance_insurance, data_income_securities, data_cash_securities, data_balance_securities, data_market, data_macro, data_com)

    # 拼接特征与label，将数据归一化

    training_data_bs = []
    one_bs = ones((len(data_bs), 1))
    data = ((data_bs - one_bs * np.mean(data_bs,axis=0)) / ( one_bs * np.std(data_bs,axis=0))).round(decimals=4, out=None)
    for i in range(0, 27968):
      training_data_bs.append([np.array(data[i]).reshape((912, 1)), [round(labels_bs[i][0] / 1000000000, 5), labels_bs[i][1]]])

    training_data_fa = []
    one_fa = ones((len(data_fa), 1))
    data = ((data_fa - one_fa * np.mean(data_fa,axis=0)) / ( one_fa * np.std(data_fa,axis=0))).round(decimals=4, out=None)
    for i in range(0, 28448):
      training_data_fa.append([np.array(data[i]).reshape((952, 1)), [round(labels_fa[i][0] / 1000000000, 5), labels_fa[i][1]]])

    return training_data_bs, training_data_fa, answer_data_bs, answer_data_fa, symbol_bs, symbol_fa


def load_xlsx():
    wb = load_workbook('xlsx/basic_data.xlsx')

    sheet_income = load_workbook('xlsx/利润表.xlsx')['Sheet']
    sheet_cash = load_workbook('xlsx/现金流量表.xlsx')['Sheet']
    sheet_balance = load_workbook('xlsx/资产负债表.xlsx')['Sheet']

    wb_income = load_workbook('xlsx/利润表_金融.xlsx')
    wb_cash = load_workbook('xlsx/现金流量表_金融.xlsx')
    wb_balance = load_workbook('xlsx/资产负债表_金融.xlsx')

    sheet_income_bank = wb_income['Bank']
    sheet_cash_bank = wb_cash['Bank']
    sheet_balance_bank = wb_balance['Bank']

    sheet_income_insurance = wb_income['Insurance']
    sheet_cash_insurance = wb_cash['Insurance']
    sheet_balance_insurance = wb_balance['Insurance']

    sheet_income_securities = wb_income['Securities']
    sheet_cash_securities = wb_cash['Securities']
    sheet_balance_securities = wb_balance['Securities']

    return wb, sheet_income, sheet_cash, sheet_balance, wb_income, wb_cash, wb_balance, sheet_income_bank, sheet_cash_bank, sheet_balance_bank, sheet_income_insurance, sheet_cash_insurance, sheet_balance_insurance, sheet_income_securities, sheet_cash_securities, sheet_balance_securities

def load_data_wrapper(data_income, data_cash, data_balance, data_income_bank, data_cash_bank, data_balance_bank, data_income_insurance, data_cash_insurance, data_balance_insurance, data_income_securities, data_cash_securities, data_balance_securities, data_market, data_macro, data_com):
    # 返回数据集
    attributes_bs = ['T_REVENUE' ,'REVENUE' ,'T_COGS' ,'COGS' ,'BIZ_TAX_SURCHG' ,'SELL_EXP' ,'ADMIN_EXP' ,'FINAN_EXP' ,'ASSETS_IMPAIR_LOSS' ,'INVEST_INCOME' ,'A_J_INVEST_INCOME' ,'OPERATE_PROFIT' ,'NOPERATE_INCOME' ,'NOPERATE_EXP' ,'NCA_DISPLOSS' ,'T_PROFIT' ,'INCOME_TAX' ,'N_INCOME' ,'N_INCOME_ATTR_P' ,'MINORITY_GAIN' ,'BASIC_EPS' ,'DILUTED_EPS' ,'OTH_COMPR_INCOME' ,'T_COMPR_INCOME' ,'COMPR_INC_ATTR_P' ,'COMPR_INC_ATTR_M_S' ,'C_FR_SALE_G_S' ,'REFUND_OF_TAX' ,'C_FR_OTH_OPERATE_A' ,'C_INF_FR_OPERATE_A' ,'C_PAID_G_S' ,'C_PAID_TO_FOR_EMPL' ,'C_PAID_FOR_TAXES' ,'C_PAID_FOR_OTH_OP_A' ,'C_OUTF_OPERATE_A' ,'N_CF_OPERATE_A' ,'PROC_SELL_INVEST' ,'GAIN_INVEST' ,'DISP_FIX_ASSETS_OTH' ,'C_FR_OTH_INVEST_A' ,'C_INF_FR_INVEST_A' ,'PUR_FIX_ASSETS_OTH' ,'C_PAID_INVEST' ,'C_OUTF_FR_INVEST_A' ,'N_CF_FR_INVEST_A' ,'C_FR_CAP_CONTR' ,'C_FR_BORR' ,'C_FR_OTH_FINAN_A' ,'C_INF_FR_FINAN_A' ,'C_PAID_FOR_DEBTS' ,'C_PAID_DIV_PROF_INT' ,'C_PAID_OTH_FINAN_A' ,'C_OUTF_FR_FINAN_A' ,'N_CF_FR_FINAN_A' ,'FOREX_EFFECTS' ,'N_CHANGE_IN_CASH' ,'N_CE_BEG_BAL' ,'N_CE_END_BAL' ,'FISCAL_PERIOD' ,'CASH_C_EQUIV' ,'NOTES_RECEIV' ,'AR' ,'PREPAYMENT' ,'OTH_RECEIV' ,'INVENTORIES' ,'OTH_CA' ,'T_CA' ,'AVAIL_FOR_SALE_FA' ,'LT_EQUITY_INVEST' ,'INVEST_REAL_ESTATE' ,'FIXED_ASSETS' ,'CIP' ,'INTAN_ASSETS' ,'GOODWILL' ,'LT_AMOR_EXP' ,'DEFER_TAX_ASSETS' ,'OTH_NCA' ,'T_NCA' ,'T_ASSETS' ,'ST_BORR' ,'NOTES_PAYABLE' ,'AP' ,'ADVANCE_RECEIPTS' ,'PAYROLL_PAYABLE' ,'TAXES_PAYABLE' ,'INT_PAYABLE' ,'DIV_PAYABLE' ,'OTH_PAYABLE' ,'NCL_WITHIN_1Y' ,'OTH_CL' ,'T_CL' ,'LT_BORR' ,'DEFER_REVENUE' ,'DEFER_TAX_LIAB' ,'OTH_NCL' ,'T_NCL' ,'T_LIAB' ,'PAID_IN_CAPITAL' ,'CAPITAL_RESER' ,'SURPLUS_RESER' ,'RETAINED_EARNINGS' ,'T_EQUITY_ATTR_P' ,'MINORITY_INT' ,'T_SH_EQUITY' ,'T_LIAB_EQUITY', 'MARKET_VALUE_1', 'MARKET_VALUE_2', 'MARKET_VALUE_3', 'MY_MACRO_DATA_A_1', 'MY_MACRO_DATA_A_2', 'MY_MACRO_DATA_A_3', 'MY_MACRO_DATA_B_1', 'MY_MACRO_DATA_B_2', 'MY_MACRO_DATA_B_3']
    attributes_fa = ['N_INCOME', 'OPERATE_PROFIT', 'AOR', 'OTH_OPER_COSTS', 'AOC', 'DILUTED_EPS', 'INVEST_INCOME', 'FOREX_GAIN', 'COMMIS_EXP', 'N_COMMIS_INCOME', 'T_COMPR_INCOME', 'COMPR_INC_ATTR_P', 'GENL_ADMIN_EXP', 'ASSETS_IMPAIR_LOSS', 'PREM_EARNED', 'N_INT_INCOME', 'T_PROFIT', 'OTH_OPER_REV', 'A_J_INVEST_INCOME', 'NOPERATE_EXP', 'N_INCOME_ATTR_P', 'MINORITY_GAIN', 'INCOME_TAX', 'NOPERATE_INCOME', 'COMPR_INC_ATTR_M_S', 'SPEC_OR', 'GROSS_PREM_WRIT', 'OTH_COMPR_INCOME', 'REVENUE', 'SPEC_OC', 'BASIC_EPS', 'F_VALUE_CHG_GAIN', 'BIZ_TAX_SURCHG', 'COGS', 'PUR_FIX_ASSETS_OTH', 'C_INF_FR_FINAN_A', 'C_PAID_FOR_DEBTS', 'C_OUTF_OPERATE_A', 'N_CHANGE_IN_CASH', 'C_FR_OTH_OPERATE_A', 'FOREX_EFFECTS', 'C_OUTF_FR_FINAN_A', 'DISP_FIX_ASSETS_OTH', 'SPEC_OCOF', 'N_CF_FR_INVEST_A', 'C_FR_ISSUE_BOND', 'C_PAID_FOR_TAXES', 'C_FR_CAP_CONTR', 'C_INF_FR_INVEST_A', 'C_PAID_INVEST', 'C_FR_OTH_INVEST_A', 'N_CF_FR_FINAN_A', 'N_DEPOS_INCR_C_FI', 'IFC_CASH_INCR', 'C_PAID_TO_FOR_EMPL', 'C_FR_MINO_S_SUBS', 'PROC_SELL_INVEST', 'GAIN_INVEST', 'C_PAID_FOR_OTH_OP_A', 'SPEC_OCIF', 'C_PAID_IFC', 'C_OUTF_FR_INVEST_A', 'C_FR_BORR', 'C_INF_FR_OPERATE_A', 'N_CE_END_BAL', 'N_CF_OPERATE_A', 'NET_INCR_DEPOS_IN_FI', 'N_CE_BEG_BAL', 'C_PAID_DIV_PROF_INT', 'RETAINED_EARNINGS', 'AE', 'OTH_LIAB', 'FIXED_ASSETS', 'T_EQUITY_ATTR_P', 'DERIV_LIAB', 'INT_RECEIV', 'SURPLUS_RESER', 'ST_BORR', 'DERIV_ASSETS', 'OTH_COMPRE_INCOME', 'LT_EQUITY_INVEST', 'ESTIMATED_LIAB', 'PUR_RESALE_FA', 'INVEST_REAL_ESTATE', 'TAXES_PAYABLE', 'T_LIAB_EQUITY', 'FOREX_DIFFER', 'CASH_C_EQUIV', 'TRADING_FA', 'OTH_ASSETS', 'BOND_PAYABLE', 'T_LIAB', 'PAYROLL_PAYABLE', 'T_ASSETS', 'PAID_IN_CAPITAL', 'LOAN_FR_OTH_BANK_FI', 'SOLD_FOR_REPUR_FA', 'LE', 'CAPITAL_RESER', 'INTAN_ASSETS', 'AVAIL_FOR_SALE_FA', 'LT_BORR', 'DEFER_TAX_LIAB', 'DEFER_TAX_ASSETS', 'TRADING_FL', 'HTM_INVEST', 'T_SH_EQUITY', 'MINORITY_INT', 'INT_PAYABLE', 'ORDIN_RISK_RESER', 'MARKET_VALUE_1', 'MARKET_VALUE_2', 'MARKET_VALUE_3', 'MY_MACRO_DATA_A_1', 'MY_MACRO_DATA_A_2', 'MY_MACRO_DATA_A_3', 'MY_MACRO_DATA_B_1', 'MY_MACRO_DATA_B_2', 'MY_MACRO_DATA_B_3']

    symbol_bs = list(data_cash.keys())
    symbol_bank = list(data_cash_bank.keys())
    symbol_insurance = list(data_income_insurance.keys())
    symbol_securities = list(data_income_securities.keys())

    symbolLenBs = len(symbol_bs)
    attrLenBs = len(attributes_bs)
    data_bs = zeros((symbolLenBs * 8, attrLenBs * 8))
    labels_bs = zeros((symbolLenBs * 8, 2))
    answer_data_bs = zeros((symbolLenBs, attrLenBs * 8))

    symbol_fa = symbol_bs.copy()
    symbol_fa.extend(symbol_bank)
    symbol_fa.extend(symbol_insurance)
    symbol_fa.extend(symbol_securities)
    symbol_fa = list(set(symbol_fa))

    symbolLenFa = len(symbol_fa)
    attrLenFa = len(attributes_fa)
    data_fa = zeros((symbolLenFa * 8, attrLenFa * 8))
    labels_fa = zeros((symbolLenFa * 8, 2))
    answer_data_fa = zeros((symbolLenFa, attrLenFa * 8))

    mapsForBs = [data_income, data_cash, data_balance, data_market, data_macro, data_com]
    mapsForFa = [data_income, data_cash, data_balance, data_income_bank, data_cash_bank, data_balance_bank, data_income_insurance, data_cash_insurance, data_balance_insurance, data_income_securities, data_cash_securities, data_balance_securities, data_market, data_macro, data_com]
    mapsForYear = [
        ['2008S1', '2008Q3', '2008A', '2009Q1', '2009S1', '2009Q3', '2009A', '2010Q1', '2010S1'],
        ['2009S1', '2009Q3', '2009A', '2010Q1', '2010S1', '2010Q3', '2010A', '2011Q1', '2011S1'],
        ['2010S1', '2010Q3', '2010A', '2011Q1', '2011S1', '2011Q3', '2011A', '2012Q1', '2012S1'],
        ['2011S1', '2011Q3', '2011A', '2012Q1', '2012S1', '2012Q3', '2012A', '2013Q1', '2013S1'],
        ['2012S1', '2012Q3', '2012A', '2013Q1', '2013S1', '2013Q3', '2013A', '2014Q1', '2014S1'],
        ['2013S1', '2013Q3', '2013A', '2014Q1', '2014S1', '2014Q3', '2014A', '2015Q1', '2015S1'],
        ['2014S1', '2014Q3', '2014A', '2015Q1', '2015S1', '2015Q3', '2015A', '2016Q1', '2016S1'],
        ['2015S1', '2015Q3', '2015A', '2016Q1', '2016S1', '2016Q3', '2016A', '2017Q1', '2017S1']
    ]
    mapsForYearAnswer = ['2016S1', '2016Q3', '2016A', '2017Q1', '2017S1', '2017Q3', '2017A', '2018Q1']
    differents = [
        [mapsForBs, symbol_bs, symbolLenBs, attributes_bs, attrLenBs, data_bs, labels_bs],
        [mapsForFa, symbol_fa, symbolLenFa, attributes_fa, attrLenFa, data_fa, labels_fa],
    ]
    a_differents = [
        [mapsForBs, symbol_bs, symbolLenBs, attributes_bs, attrLenBs, answer_data_bs],
        [mapsForFa, symbol_fa, symbolLenFa, attributes_fa, attrLenFa, answer_data_fa]
    ]

    for different in differents:
        for dictData in different[0]:
            for s in range(0, different[2]):
                num = different[1][s]
                for i in range(0, 8):
                    for j in range(0, 8):
                        for a in range(0, different[4]):
                            year = mapsForYear[i][j]
                            if (num in dictData and year in dictData[num]):
                                attr = different[3][a]
                                if (attr in dictData[num][year]):
                                    row = (s * 8) + i
                                    col = j * different[4] + a
                                    different[5][row][col] = float(dictData[num][year][attr])

                    if (num in dictData and mapsForYear[i][8] in dictData[num]):
                        if ('REVENUE' in dictData[num][mapsForYear[i][8]]):
                            row = (s * 8) + i
                            different[6][row][0] = float(dictData[num][mapsForYear[i][8]]['REVENUE'])

                    if (num in dictData and '2018S1' in dictData[num]):
                        if ('MARKET_VALUE_2' in dictData[num]['2018S1']):
                            row = (s * 8) + i
                            different[6][row][1] = round(float(dictData[num]['2018S1']['MARKET_VALUE_2']) / 100000000, 2)

    for different in a_differents:
        for dictData in different[0]:
            for s in range(0, different[2]):
                num = different[1][s]
                for i in range(0, 8):
                    for a in range(0, different[4]):
                        year = mapsForYearAnswer[i]
                        if (num in dictData and year in dictData[num]):
                            attr = different[3][a]
                            if (attr in dictData[num][year]):
                                row = s
                                col = i * different[4] + a
                                different[5][row][col] = float(dictData[num][year][attr])

    return data_bs, labels_bs, answer_data_bs, data_fa, labels_fa, answer_data_fa, symbol_bs, symbol_fa


def load_data(sheet_income, sheet_cash, sheet_balance, sheet_income_bank, sheet_cash_bank, sheet_balance_bank, sheet_income_insurance, sheet_cash_insurance, sheet_balance_insurance, sheet_income_securities, sheet_cash_securities, sheet_balance_securities):
    # 将一般工商业3张表数据及basic data合成一张表
    mapsForSheet = [sheet_income, sheet_cash, sheet_balance, sheet_income_bank, sheet_cash_bank, sheet_balance_bank, sheet_income_insurance, sheet_cash_insurance, sheet_balance_insurance, sheet_income_securities, sheet_cash_securities, sheet_balance_securities]
    test = []
    for i in range(0, 12):
        test.append(handle_financial_data(mapsForSheet[i]))
    return test[0], test[1], test[2], test[3], test[4], test[5], test[6], test[7], test[8], test[9], test[10], test[11]

def load_basic_data(wb):
    data_market, mapsForSymbol = handle_basic_data_for_market(wb)
    data_macro = handle_basic_data_for_macro(wb, mapsForSymbol)
    data_com = handle_basic_data_for_com(wb)
    return data_market, data_macro, data_com

def handle_financial_data(sheet):
    # 处理一般工商业、金融(银行、证券、保险)共3*4=12张表数据
    rows = sheet.max_row
    column = sheet.max_column
    data = {}     # 以股票代码为key
    quarter = {}  # 以各季度为key
    symbol = sheet.cell(row = 2, column = 1).value
    string = []
    for i in range(1, column + 1):
        string.append(sheet.cell(row = 1, column = i).value)
    for i in range(2, rows + 1):
        if (sheet.cell(row = i, column = 1).value != symbol):
            data[symbol] = quarter
            quarter = {}
            symbol = sheet.cell(row = i, column = 1).value
        qKey = ''.join([sheet.cell(row = i, column = 2).value[0:4],sheet.cell(row = i, column = 3).value])
        if (not qKey in quarter):
            quarter[qKey] = {}
            for j in range(4, column + 1):
                value = sheet.cell(row = i, column = j).value
                if (value != None):
                    quarter[qKey][string[j-1]] = value
                else:
                    quarter[qKey][string[j-1]] = 0
        else:
            newDict = {}
            for j in range(4, column + 1):
                value = sheet.cell(row = i, column = j).value
                if (value != None):
                    newDict[string[j-1]] = value
            quarter[qKey].update(newDict)
    data[symbol] = quarter
    return data

def handle_basic_data_for_market(wb):
    # 处理market_data
    sheet = wb['market_data']
    # rows = 333139
    rows = sheet.max_row
    data = {}     # 以股票代码为key
    quarter = {}  # 以各季度为key
    date = {}     # 存储一个年的数据
    mapsForSymbol = {}    # 行业
    symbol = sheet.cell(row = 2, column = 1).value
    year = sheet.cell(row = 2, column = 2).value.year
    mapsForSymbol[sheet.cell(row = 2, column = 5).value] = [symbol]
    strings = ['Q1', 'S1', 'Q3', 'A']
    for i in range(2, rows + 2):
        if (sheet.cell(row = i, column = 1).value != symbol or sheet.cell(row = i, column = 2).value.year != year):
            maps = {
                1: {'q': str(year) + 'Q1', 'v': 'MARKET_VALUE_1'},
                2: {'q': str(year) + 'Q1', 'v': 'MARKET_VALUE_2'},
                3: {'q': str(year) + 'Q1', 'v': 'MARKET_VALUE_3'},
                4: {'q': str(year) + 'S1', 'v': 'MARKET_VALUE_1'},
                5: {'q': str(year) + 'S1', 'v': 'MARKET_VALUE_2'},
                6: {'q': str(year) + 'S1', 'v': 'MARKET_VALUE_3'},
                7: {'q': str(year) + 'Q3', 'v': 'MARKET_VALUE_1'},
                8: {'q': str(year) + 'Q3', 'v': 'MARKET_VALUE_2'},
                9: {'q': str(year) + 'Q3', 'v': 'MARKET_VALUE_3'},
                10: {'q': str(year) + 'A', 'v': 'MARKET_VALUE_1'},
                11: {'q': str(year) + 'A', 'v': 'MARKET_VALUE_2'},
                12: {'q': str(year) + 'A', 'v': 'MARKET_VALUE_3'}
            }
            for qKey in strings:
                quarter[str(year) + qKey] = {
                    'MARKET_VALUE_1': 0,
                    'MARKET_VALUE_2': 0,
                    'MARKET_VALUE_3': 0
                }
            for month in date:
                quarter[maps[month]['q']][maps[month]['v']] = date[month]
            date = {}
            if (i != rows + 1):
                year = sheet.cell(row = i, column = 2).value.year
            if (i != rows + 1 and sheet.cell(row = i, column = 1).value != symbol):
                data[symbol] = quarter
                quarter = {}
                symbol = sheet.cell(row = i, column = 1).value
                if (sheet.cell(row = i, column = 5).value in mapsForSymbol):
                    mapsForSymbol[sheet.cell(row = i, column = 5).value].append(symbol)
                else:
                    mapsForSymbol[sheet.cell(row = i, column = 5).value] = [symbol]
        if (i != rows + 1):
            date[sheet.cell(row = i, column = 2).value.month] = sheet.cell(row = i, column = 3).value
    data[symbol] = quarter
    return data, mapsForSymbol

def handle_basic_data_for_macro(wb, mapsForSymbol):
    # 处理macro_name
    sheet = wb['macro_data']
    # rows = 28895
    rows = sheet.max_row
    data = {}
    quarter = {}  # 以各季度为key
    date = {}     # 存储一个年的数据
    indic = sheet.cell(row = 2, column = 1).value
    year = sheet.cell(row = 2, column = 3).value.year
    strings = ['Q1', 'S1', 'Q3', 'A']
    mapsForType = {
        '1020000004': ['Non-Ferrous Metals', 'Mechanical Equipment', 'Chemical Industry'],
        '1020000008': ['Non-Ferrous Metals', 'Chemical Industry'],
        '1020001544': ['Household Appliances', 'Electronic Equipment'],
        '1030000014': ['Textile and Garment', 'Light Manufacturing'],
        '1030000016': ['Building Decorations', 'Building Material'],
        '1030000018': ['Textile and Garment', 'Light Manufacturing'],
        '1030000020': ['Building Decorations', 'Building Material'],
        '1040000046': ['Medicine and Biology', 'Food and Beverage', 'Media'],
        '1040000050': ['Medicine and Biology', 'Food and Beverage', 'Media'],
        '1040002190': ['Defense and Military'],
        '1050000027': ['Real Estate'],
        '1070000035': ['Financial Service'],
        '1070000039': ['Financial Service'],
        '1080000235': ['Composite'],
        '1090000363': ['Bank'],
        '1090000365': ['Bank'],
        '1100000874': ['Electronics', 'Information Devices', 'Computer', 'Communication'],
        '1100002293': ['Steel'],
        '1100005542': ['Composite'],
        '1170000598': ['Information Services'],
        '1170000618': ['Information Services'],
        '1170000641': ['Commercial Trade'],
        '1170003826': ['Transportation'],
        '1170003828': ['Building Materials', 'Delivery Equipment'],
        '1170003830': ['Transportation'],
        '1170004261': ['Building Materials', 'Delivery Equipment'],
        '1170004326': ['Mechanical Equipment'],
        '1170007422': ['Commercial Trade'],
        '2020000719': ['Defense and Military'],
        '2020100020': ['Steel', 'Mining'],
        '2020100024': ['Mining'],
        '2020101521': ['Utilities', 'Household Appliances', 'Electronic Equipment'],
        '2020101526': ['Utilities'],
        '2070100748': ['Automobile'],
        '2070100795': ['Automobile'],
        '2070109977': ['Leisure Service'],
        '2070113040': ['Leisure Service'],
        '2090100464': ['Electronics', 'Information Devices', 'Computer', 'Communication'],
        '2160000004': ['Non-bank Finance'],
        '2160001002': ['Animal Husbandry and Fishery'],
        '2160001011': ['Animal Husbandry and Fishery'],
        '2170002035': ['Real Estate'],
        '2210200588': ['Non-bank Finance'],
    }
    mapsForValue = {
        '1020000004': 2,
        '1020000008': 1,
        '1020001544': 1995,
        '1030000014': 29,
        '1030000016': 43.4,
        '1030000018': 32.2,
        '1030000020': 39.5,
        '1040000046': 1,
        '1040000050': 1,
        '1040002190': 42,
        '1050000027': 2265,
        '1070000035': 543790,
        '1070000039': 1676800,
        '1080000235': 6,
        '1090000363': 4.9,
        '1090000365': 3.25,
        '1100000874': 28,
        '1100002293': 217,
        '1100005542': 94477,
        '1170000598': 1866,
        '1170000618': 21560.73,
        '1170000641': 79540.01,
        '1170003826': 1007.94,
        '1170003828': 1134.39,
        '1170003830': 821.62,
        '1170004261': 3440.67,
        '1170004326': 1194.87,
        '1170007422': 5,
        '2020000719': 27.79,
        '2020100020': 17000.79,
        '2020100024': 13965,
        '2020101521': 3134,
        '2020101526': 2324,
        '2070100748': 44,
        '2070100795': 46,
        '2070109977': 461378,
        '2070113040': 468973,
        '2090100464': 5.1,
        '2160000004': 290,
        '2160001002': 2385,
        '2160001011': 8,
        '2170002035': 27194,
        '2210200588': 6.66
    }
    for i in range(2, rows + 2):
        if (i != rows + 1 and not str(sheet.cell(row = i, column = 1).value) in mapsForType):
            continue
        if (sheet.cell(row = i, column = 1).value != indic or sheet.cell(row = i, column = 3).value.year != year):
            mapsForKey = {
                1: {'q': str(year) + 'Q1', 'v': 'DATA_1'},
                2: {'q': str(year) + 'Q1', 'v': 'DATA_2'},
                3: {'q': str(year) + 'Q1', 'v': 'DATA_3'},
                4: {'q': str(year) + 'S1', 'v': 'DATA_1'},
                5: {'q': str(year) + 'S1', 'v': 'DATA_2'},
                6: {'q': str(year) + 'S1', 'v': 'DATA_3'},
                7: {'q': str(year) + 'Q3', 'v': 'DATA_1'},
                8: {'q': str(year) + 'Q3', 'v': 'DATA_2'},
                9: {'q': str(year) + 'Q3', 'v': 'DATA_3'},
                10: {'q': str(year) + 'A', 'v': 'DATA_1'},
                11: {'q': str(year) + 'A', 'v': 'DATA_2'},
                12: {'q': str(year) + 'A', 'v': 'DATA_3'}
            }
            for qKey in strings:
                if (not (str(year) + qKey) in quarter):
                    quarter[str(year) + qKey] = {}
            for month in date:
                quarter[mapsForKey[month]['q']][mapsForKey[month]['v']] = round(date[month] / mapsForValue[str(indic)], 2)
            date = {}
            if (i != rows + 1):
                year = sheet.cell(row = i, column = 3).value.year
            if (i != rows + 1 and sheet.cell(row = i, column = 1).value != indic):
                for types in mapsForType[str(indic)]:
                    for symbol in mapsForSymbol[types]:
                        if(not symbol in data):
                            data[symbol] = {}
                        for qKey in quarter:
                            if(qKey in data[symbol]):
                                data[symbol][qKey].update({
                                   'MY_MACRO_DATA_B_1': quarter[qKey]['DATA_1'] if 'DATA_1' in quarter[qKey] else 0,
                                   'MY_MACRO_DATA_B_2': quarter[qKey]['DATA_2'] if 'DATA_2' in quarter[qKey] else 0,
                                   'MY_MACRO_DATA_B_3': quarter[qKey]['DATA_3'] if 'DATA_3' in quarter[qKey] else 0
                               })
                            else:
                               data[symbol][qKey] = {
                                   'MY_MACRO_DATA_A_1': quarter[qKey]['DATA_1'] if 'DATA_1' in quarter[qKey] else 0,
                                   'MY_MACRO_DATA_A_2': quarter[qKey]['DATA_2'] if 'DATA_2' in quarter[qKey] else 0,
                                   'MY_MACRO_DATA_A_3': quarter[qKey]['DATA_3'] if 'DATA_3' in quarter[qKey] else 0
                               }
                quarter = {}
                indic = sheet.cell(row = i, column = 1).value
        if (i != rows + 1):
            date[sheet.cell(row = i, column = 3).value.month] = sheet.cell(row = i, column = 4).value
    return data

def handle_basic_data_for_com(wb):
    # 处理company_operating
    sheet = wb['company_operating']
    rows = sheet.max_row
    data = {}     # 以股票代码为key
    quarter = {}  # 以各季度为key
    date = {}     # 存储一个年的数据
    symbol = sheet.cell(row = 2, column = 1).value
    year = sheet.cell(row = 2, column = 3).value.year
    strings = ['Q1', 'S1', 'Q3', 'A']
    mapsForValue = {
        '600004': [22241, 2645606],
        '600115': [2569],
        '600269': [1074, 24329],
        '601333': [40866000, 1352000],
        '600029': [4999],
        '600018': [2157, 153],
        '600009': [20179, 10244],
        '000089': [177],
        '601111': [1000, 1000],
        '601018': [3023, 109],
        '601231': [900000000],
        '600221': [214577],
        '600383': [37000],
        '601088': [7427],
        '600048': [77500],
        '601898': [1515, 5350000],
        '000536': [280000000]
    }
    for i in range(2, rows + 2):
        if (i != rows + 1 and sheet.cell(row = i, column = 3).value.year < 2008):
            year = sheet.cell(row = 2, column = 3).value.year
            continue
        if (sheet.cell(row = i, column = 1).value != symbol or sheet.cell(row = i, column = 3).value.year != year):
            mapsForKey = {
                1: {'q': str(year) + 'Q1', 'v1': 'MY_MACRO_DATA_A_1', 'v2': 'MY_MACRO_DATA_A_1'},
                2: {'q': str(year) + 'Q1', 'v1': 'MY_MACRO_DATA_A_2', 'v2': 'MY_MACRO_DATA_B_2'},
                3: {'q': str(year) + 'Q1', 'v1': 'MY_MACRO_DATA_A_3', 'v2': 'MY_MACRO_DATA_B_3'},
                4: {'q': str(year) + 'S1', 'v1': 'MY_MACRO_DATA_A_1', 'v2': 'MY_MACRO_DATA_B_1'},
                5: {'q': str(year) + 'S1', 'v1': 'MY_MACRO_DATA_A_2', 'v2': 'MY_MACRO_DATA_B_2'},
                6: {'q': str(year) + 'S1', 'v1': 'MY_MACRO_DATA_A_3', 'v2': 'MY_MACRO_DATA_B_3'},
                7: {'q': str(year) + 'Q3', 'v1': 'MY_MACRO_DATA_A_1', 'v2': 'MY_MACRO_DATA_B_1'},
                8: {'q': str(year) + 'Q3', 'v1': 'MY_MACRO_DATA_A_2', 'v2': 'MY_MACRO_DATA_B_2'},
                9: {'q': str(year) + 'Q3', 'v1': 'MY_MACRO_DATA_A_3', 'v2': 'MY_MACRO_DATA_B_3'},
                10: {'q': str(year) + 'A', 'v1': 'MY_MACRO_DATA_A_1', 'v2': 'MY_MACRO_DATA_B_1'},
                11: {'q': str(year) + 'A', 'v1': 'MY_MACRO_DATA_A_2', 'v2': 'MY_MACRO_DATA_B_2'},
                12: {'q': str(year) + 'A', 'v1': 'MY_MACRO_DATA_A_3', 'v2': 'MY_MACRO_DATA_B_3'}
            }
            for qKey in strings:
                if (not (str(year) + qKey) in quarter):
                    quarter[str(year) + qKey] = {}
            for month in date:
                if (mapsForKey[month]['v1'] in quarter[mapsForKey[month]['q']]):
                    quarter[mapsForKey[month]['q']][mapsForKey[month]['v2']] = round(date[month] / mapsForValue[str(symbol)][1], 2)
                else:
                    quarter[mapsForKey[month]['q']][mapsForKey[month]['v1']] = round(date[month] / mapsForValue[str(symbol)][0], 2)
            date = {}
            if (i != rows + 1):
                year = sheet.cell(row = i, column = 3).value.year
            if (i != rows + 1 and sheet.cell(row = i, column = 1).value != symbol):
                data[symbol] = quarter
                quarter = {}
                symbol = sheet.cell(row = i, column = 1).value
        if (i != rows + 1):
            date[sheet.cell(row = i, column = 3).value.month] = sheet.cell(row = i, column = 4).value
    data[symbol] = quarter
    return data